import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
PLANILHA = BASE_DIR / "INDICADORES DO SUPORTE DE 2026.xlsx"
EXPORT_PADRAO = BASE_DIR / "comparativo_exportado.xlsx"

MESES = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARCO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}


def normalizar(valor):
    texto = "" if valor is None else str(valor)
    trocas = {
        "Á": "A", "À": "A", "Â": "A", "Ã": "A",
        "É": "E", "Ê": "E",
        "Í": "I",
        "Ó": "O", "Ô": "O", "Õ": "O",
        "Ú": "U",
        "Ç": "C",
    }
    texto = texto.upper()
    for original, novo in trocas.items():
        texto = texto.replace(original, novo)
    return texto


def escolher_aba(workbook):
    hoje = datetime.now()
    mes = MESES[hoje.month]
    ano = str(hoje.year)[-2:]

    for nome in workbook.sheetnames:
        nome_normalizado = normalizar(nome).replace(".", " ")
        if mes in nome_normalizado and ano in nome_normalizado:
            return workbook[nome]

    for nome in reversed(workbook.sheetnames):
        aba = workbook[nome]
        if "METRICA MATRIZ" in normalizar(aba["A1"].value):
            return aba

    raise ValueError("Aba mensal nao encontrada.")


def numero(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if hasattr(valor, "hour") and hasattr(valor, "minute") and hasattr(valor, "second"):
        return (valor.hour * 3600 + valor.minute * 60 + valor.second) / 86400

    texto = str(valor).strip().replace(",", ".")
    if ":" in texto:
        partes = [int(parte) for parte in texto.replace("::", ":").split(":") if parte]
        if len(partes) == 3:
            return (partes[0] * 3600 + partes[1] * 60 + partes[2]) / 86400
        if len(partes) == 2:
            return (partes[0] * 60 + partes[1]) / 86400

    try:
        return float(texto)
    except ValueError:
        return None


def valor_preenchido(valor):
    return numero(valor) is not None


def colunas_semanas(aba):
    semanas_mes = []
    semanas_fallback = []

    for coluna in range(2, aba.max_column + 1):
        rotulo = normalizar(aba.cell(row=1, column=coluna).value)
        if "SEMANA" not in rotulo:
            continue

        tem_dados = all(
            valor_preenchido(aba.cell(row=linha, column=coluna).value)
            for linha in (2, 3, 6)
        )
        if not tem_dados:
            continue

        semanas_fallback.append(coluna)
        if "ULTIMA" not in rotulo:
            semanas_mes.append(coluna)

    semanas = semanas_mes if len(semanas_mes) >= 2 else semanas_fallback
    if len(semanas) < 2:
        raise ValueError("Nao existem duas semanas preenchidas para comparacao.")
    return semanas[-2], semanas[-1]


def segundos(valor):
    convertido = numero(valor)
    if convertido is None:
        raise ValueError("Tempo invalido na planilha.")
    return int(round(convertido * 86400))


def formatar_segundos(valor):
    valor = int(round(valor))
    horas = valor // 3600
    minutos = (valor % 3600) // 60
    seg = valor % 60
    if horas:
        return f"{horas}h{minutos:02d}m{seg:02d}s"
    return f"{minutos}m{seg:02d}s"


def formatar_segundos_com_sinal(valor):
    if valor == 0:
        return "0m00s"
    sinal = "-" if valor < 0 else "+"
    return sinal + formatar_segundos(abs(valor))


def formatar_decimal(valor):
    return f"{valor:.2f}".replace(".", ",")


def formatar_decimal_com_sinal(valor):
    if valor == 0:
        return "0,00"
    sinal = "+" if valor > 0 else "-"
    return sinal + formatar_decimal(abs(valor))


def card_tempo(indicador, nome, anterior, atual, meta_segundos):
    anterior_seg = segundos(anterior)
    atual_seg = segundos(atual)
    diferenca = atual_seg - anterior_seg
    dentro = atual_seg <= meta_segundos

    return {
        "indicador": indicador,
        "nome": nome,
        "meta": "até " + formatar_segundos(meta_segundos),
        "semana_passada": formatar_segundos(anterior_seg),
        "semana_atual": formatar_segundos(atual_seg),
        "diferenca": formatar_segundos_com_sinal(diferenca),
        "status": "Dentro da meta" if dentro else "Fora da meta",
        "status_classe": "dentro" if dentro else "fora",
        "diferenca_classe": "melhor" if diferenca < 0 else ("pior" if diferenca > 0 else "neutro"),
    }


def card_csat(anterior, atual):
    anterior_num = numero(anterior)
    atual_num = numero(atual)
    if anterior_num is None or atual_num is None:
        raise ValueError("CSAT invalido na planilha.")

    diferenca = atual_num - anterior_num
    dentro = atual_num >= 4.5

    return {
        "indicador": "CSAT",
        "nome": "Qualidade Percebida na Avaliação Geral",
        "meta": "4,50",
        "semana_passada": formatar_decimal(anterior_num),
        "semana_atual": formatar_decimal(atual_num),
        "diferenca": formatar_decimal_com_sinal(diferenca),
        "status": "Dentro da meta" if dentro else "Fora da meta",
        "status_classe": "dentro" if dentro else "fora",
        "diferenca_classe": "melhor" if diferenca > 0 else ("pior" if diferenca < 0 else "neutro"),
    }


def calcular():
    if not PLANILHA.exists():
        raise FileNotFoundError("Planilha INDICADORES DO SUPORTE DE 2026.xlsx nao encontrada.")

    workbook = load_workbook(PLANILHA, data_only=True)
    aba = escolher_aba(workbook)
    semana_passada, semana_atual = colunas_semanas(aba)

    return [
        card_tempo(
            "TMA",
            "Tempo Médio de Atendimento",
            aba.cell(row=2, column=semana_passada).value,
            aba.cell(row=2, column=semana_atual).value,
            45 * 60,
        ),
        card_tempo(
            "TMR",
            "Tempo Médio de Resposta",
            aba.cell(row=3, column=semana_passada).value,
            aba.cell(row=3, column=semana_atual).value,
            2 * 60 + 40,
        ),
        card_csat(
            aba.cell(row=6, column=semana_passada).value,
            aba.cell(row=6, column=semana_atual).value,
        ),
    ]


def exportar(cards, caminho):
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Comparativo"

    cabecalho = [
        "Indicador",
        "Meta",
        "Semana passada",
        "Semana atual",
        "Diferença",
        "Status da meta",
    ]
    aba.append(cabecalho)

    for card in cards:
        aba.append([
            card["indicador"],
            card["meta"],
            card["semana_passada"],
            card["semana_atual"],
            card["diferenca"],
            card["status"],
        ])

    azul = PatternFill("solid", fgColor="174EA6")
    for celula in aba[1]:
        celula.fill = azul
        celula.font = Font(color="FFFFFF", bold=True)
        celula.alignment = Alignment(horizontal="center")

    for largura, coluna in zip([14, 18, 20, 18, 16, 18], range(1, 7)):
        aba.column_dimensions[aba.cell(row=1, column=coluna).column_letter].width = largura

    for linha in aba.iter_rows(min_row=2):
        for celula in linha:
            celula.alignment = Alignment(horizontal="center")

    workbook.save(caminho)
    return caminho


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--export", nargs="?", const=str(EXPORT_PADRAO))
    args = parser.parse_args()

    cards = calcular()

    if args.export:
        caminho = Path(args.export).resolve()
        exportar(cards, caminho)
        print(caminho)
        return

    if args.json:
        print(json.dumps({"cards": cards}, ensure_ascii=False))
        return

    print(json.dumps({"cards": cards}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
